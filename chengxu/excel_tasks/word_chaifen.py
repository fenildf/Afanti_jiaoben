# -*-coding:utf8-*-
import os
import sys
import json
import re
import urllib
import requests
import logging
import hashlib
import pymysql
import time
import xlrd
import openpyxl
import random
from collections import OrderedDict
from sogou_xml.spider_zdic.deallib.table_common import conn_13309

# LOGGING_FORMAT = '%(asctime)-15s:%(levelname)s: %(message)s'
# logging.basicConfig(format=LOGGING_FORMAT, level=logging.INFO,
#                                filename='working/baidu_idiom_bulou.log', filemode='a')

_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))).replace('\\', '/')
INPUT_PATH = _DIR + '/inputfile/'
OUT_PATH = _DIR + '/outputfile/xlsx/'
CONFIG_FILE = os.path.join(_DIR, 'config')
output_file = u'汉字拆分.xlsx'


def dealXlsx(datas, ziduan):
    filename = OUT_PATH + output_file
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    sheet = workbook.create_sheet(ziduan, 0)
    for i in range(len(datas)):
        a = sheet.cell(row=i + 1, column=1, value=datas[i]['word_name'])
        b = sheet.cell(row=i + 1, column=2, value=datas[i]['bushou'])
        c = sheet.cell(row=i + 1, column=3, value=datas[i]['jiegou'])
        d = sheet.cell(row=i + 1, column=4, value=datas[i]['chaifen'])

    workbook.save(filename)


def read_table():
    cur = conn_13309.cursor()
    sql = "select word_name,bushou,jiegou,chaifen from word_bhl_2018"
    cur.execute(sql)
    datas = cur.fetchall()
    dealXlsx(datas,u'部首_结构_拆分')



if __name__ == '__main__':
    read_table()
    print('dada')
